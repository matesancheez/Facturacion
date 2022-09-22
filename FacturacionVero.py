from time import sleep
import openpyxl 
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from os import system  


system('mode con: cols=100 lines=30')


wb = openpyxl.Workbook()
ws = wb.active
        

def cargarFecha():
    print()
    print('Ingrese la fecha en la que realizo el tratamiento: ')
    dia = int(input('Dia >> '))
    mes = int(input('Mes >> '))
    ano = int(input('Año >> '))
    print()

    fecha = '{}\{}\{}'.format(dia, mes, ano)
    
    return fecha


def main():
    total = 2
    inicio = inicio2 = 2
    print()
    print('<'*33, '>'*33)
    print('<'*10,'Sistema de facturacion Dra. Veronica Sanchez', '>'*11)
    print('<'*33, '>'*33)
    print()
    sleep(0.3)
    mes = input("¿Que mes esta por facturar? >>> ")
    x = int(input("¿Cuantas fichas desea cargar? >>> "))


    for i in range(x):
        print()
        print('PACIENTE NUMERO {}'.format(i+1))
        nombre = input('Ingrese el NOMBRE del {} paciente >> '.format(i + 1))
        apellido = input('Ingrese el APELLIDO del {} paciente >> '.format(i + 1))
        afiliado = input('Ingrese el Numero de Afiliado del {} paciente >> '.format(i + 1))
        fecha = cargarFecha()
        plan = input('Ingrese el PLAN del {} paciente >> '.format(i + 1))
        
        nDefichas = int(input("¿Cuantos tratamientos tiene {} {}? >>> ".format(nombre, apellido)))
        total += nDefichas 

        for k in range(nDefichas):
            pos = str(inicio + k )
            ws['A' + pos] = afiliado

            ws['A' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['A' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)

            ws['B' + pos] = nombre + ' ' + apellido
            
            ws['B' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['B' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)

            ws['C' + pos] = fecha
            
            ws['C' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['C' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)

            ws['H' + pos] = plan
            
            ws['H' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['H' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)


        inicio = total 
        


        
        for j in range(nDefichas):
            pos = str(j + inicio2)
            print()
            print("<" * 25 + ">" * 25)
            print('>>>> Tratamiento numero {} del paciente {} {} '.format(j+1, nombre, apellido))
            print("<" * 25 + ">" * 25)
            print()
            
            print()
            codigo = int(input(">> Codigo del tratamiento: "))
            pieza = int(input(">>  Numero de pieza: "))
            cara =  input('>> Cara del tratamiento: ')
            honorarios = int(input('>> Cuales con los HONORARIOS del tratamiento: '))

            ws['D' + pos] = codigo
            
            ws['D' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['D' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)
            

            ws['E' + pos] = pieza
            
            ws['E' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['E' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)
            

            ws['F' + pos] = cara
            
            ws['F' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['F' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)
            

            ws['G' + pos] = honorarios
            
            ws['G' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['G' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)
            

            ws['H' + pos] = '50673'
            
            ws['H' + pos].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
            ws['H' + pos].alignment = Alignment(horizontal='center',
                                            vertical='center',)
            

        inicio2 = total
            

    ###################
    #ESTILOS DE CELDAS#
    ###################

    ws['A1'] = 'Numero de afiliado'
    
    #ESTILOS
    ws['A1'].font = Font(size=14)
    ws['A1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['A1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['A1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)
    
    ws['B1'] = 'Nombre y apellido'
    
    #ESTILOS
    ws['B1'].font = Font(size=14)
    ws['B1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['B1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['B1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)
    
    ws['C1'] = 'Fecha'
    
    #ESTILOS
    ws['C1'].font = Font(size=14)
    ws['C1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['C1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['C1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)
    
    ws['D1'] = 'Codigo'
    
    #ESTILOS
    ws['D1'].font = Font(size=14)
    ws['D1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['D1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['D1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)
    
    ws['E1'] = 'Pieza'
    
    #ESTILOS
    ws['E1'].font = Font(size=14)
    ws['E1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['E1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['E1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)
    
    ws['F1'] = 'Cara'
    
    #ESTILOS
    ws['F1'].font = Font(size=14)
    ws['F1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['F1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['F1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)
    
    ws['G1'] = 'Honorarios'
    
    #ESTILOS
    ws['G1'].font = Font(size=14)
    ws['G1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['G1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['G1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)
    
    ws['H1'] = 'Prestador'
    
    #ESTILOS
    ws['H1'].font = Font(size=14)
    ws['H1'].fill = PatternFill(fill_type='solid',
                                start_color='00CCFFCC')
    ws['H1'].border = Border(left=Side(border_style='thin',
                              color='00000000'),
                              right=Side(border_style='thin',
                              color='00000000'),
                              top=Side(border_style='thin',
                              color='00000000'),
                              bottom=Side(border_style='thin',
                              color='00000000'))
    ws['H1'].alignment = Alignment(horizontal='center',
                                    vertical='center',)


    
    #ANCHO DE COLUMNAS 

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15


    print('Generando Excel...')
    sleep(0.5)
    print('Generando Excel...')
    sleep(0.5)
    print('Excel generado correctamente. Gracias.')
    sleep(0.5)
    wb.save('facturacion_{}.xlsx'.format(mes))


if __name__ == "__main__":
    main()
